import openpyxl
from datetime import datetime

def add_data_to_excel(user,prompt,reasoning,responce,filename='prompt_responce.xlsx'):
    try:
        workbook=openpyxl.load_workbook(filename)
        print('loaded file')
    except FileNotFoundError:
        workbook=openpyxl.Workbook()
        print('file create')

    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    if 'Prompt_Responces' not in workbook.sheetnames:
        workbook.create_sheet('Prompt_Responces')

    sheet=workbook['Prompt_Responces']

    if sheet.max_row == 1 and sheet.cell(row=1,column=1).value is None:
        sheet.cell(1,1,"ID")
        sheet.cell(1,2,'User')
        sheet.cell(1,3,'Time Stamp')
        sheet.cell(1,4,'prompt')
        sheet.cell(1,5,'reasoning')
        sheet.cell(1,6,'responce')

    sheet.append([sheet.max_row,user,str(datetime.now()),prompt,reasoning,responce])

    workbook.save(filename)